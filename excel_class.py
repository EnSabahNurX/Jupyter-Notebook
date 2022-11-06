import openpyxl
import datetime
wb = openpyxl.load_workbook("02_Excel_data.xlsx")
# print(wb.sheetnames)
folha_um = wb.sheetnames[0]
# print(folha_um)
folha_dois = wb['address']
# print(folha_dois)
# print(folha_dois['A1'].value)
#folha_dois.cell(row=5, column=2).value
multiple_cells = folha_dois['A1:B3']

# for row in multiple_cells:
#     for cell in row:
#         print(cell.value)

# for fila in folha_dois.rows:
#     for cell in fila:
#         print(cell.value)

# for coluna in folha_dois.columns:
#     for cell in coluna:
#         print(cell.value)

folha_tres = wb['ventas']
# folha_tres['B6'] = 5
# folha_tres['C6'] = 6
# folha_tres.cell(row=6, column=4, value=5)
# folha_tres['E2'].value = datetime.datetime.now()
# wb.save("02_Excel_data.xlsx")
folha_tres['B6'] = '=SUM(B5,3)'
wb.save("02_Excel_data.xlsx")
folha_tres.append(['Junior', 99, 100])
wb.save("02_Excel_data.xlsx")
